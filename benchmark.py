# -*- coding: utf-8 -*-
import argparse
from functools import wraps
import re
import timeit
from itertools import cycle

ROWS = 1000
COLUMNS = 100
RUN_COUNT = 10

VALUES = cycle([1, None, "foobar", 2.32])


def skip(description):
    def skip_decorator(fun):

        @wraps(fun)
        def wrapped(*args, **kwargs):
            fun(*args, **kwargs)

        wrapped.skip = description

        return wrapped

    return skip_decorator


def get_benchmarks():
    return [
        item
        for item in globals().values()
        if callable(item) and item.__name__.startswith('benchmark')
    ]


def benchmark_xlwt():
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('A Test Sheet')

    for row, value in zip(range(ROWS), VALUES):
        for column in range(COLUMNS):
            sheet.write(row, column, value)

    workbook.save('benchmark_xlwt.xlsx')


def benchmark_openpyxl_rows():
    """OpenPyXL using sheet.append('row')

    As documentation says this should be faster:
    http://pythonhosted.org/openpyxl/optimized.html#optimized-writer
    """

    import openpyxl

    workbook = openpyxl.workbook.Workbook()
    sheet = workbook.create_sheet()
    sheet.title = 'Sheet1'
    # note: pyopenxl indexes rows and columns starting from 1
    for row, value in zip(range(1, ROWS + 1), VALUES):
        sheet.append([str(value) for _ in range(1, COLUMNS + 1)])

    workbook.save('benchmark_openpyxl_rows.xlsx')


def benchmark_openpyxl():
    """OpenPyXL using sheet.cell().value = value"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.workbook.Workbook()
    sheet = workbook.create_sheet()
    sheet.title = 'Sheet1'

    # note: pyopenxl indexes rows and columns starting from 1
    for row, value in zip(range(1, ROWS + 1), VALUES):
        for column in range(1, COLUMNS + 1):
            sheet.cell(row, column).value = value

    workbook.save('benchmark_openpyxl.xlsx')

def benchmark_pylightxl():
    import pylightxl as xl

    db = xl.Database()

    db.add_ws(ws="Sheet1")
    ws = db.ws(ws="Sheet1")

    for row, value in zip(range(1, ROWS + 1), VALUES):
        for column in range(1, COLUMNS + 1):
            ws.update_index(row=row, col=column, val=value)

    xl.writexl(db=db, fn="benchmark_pylightxl.xlsx")


def benchmark_pyexcelerate():
    import pyexcelerate

    workbook = pyexcelerate.Workbook()

    data = [
        [value for column in range(COLUMNS)]
         for __, value in zip(range(ROWS), VALUES)
    ]

    workbook.new_sheet('Test 1', data=data)
    workbook.save('benchmark_pyexcelerate.xlsx')


def benchmark_xlsxwriter():
    import xlsxwriter
    workbook = xlsxwriter.Workbook('benchmark_xlsxwriter.xlsx')
    sheet = workbook.add_worksheet()

    for row, value in zip(range(ROWS), VALUES):
        for column in range(COLUMNS):
            sheet.write(column, row, value)

    workbook.close()


def benchmark_csv():
    import csv

    with open('benchmark_scv.csv', 'w') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows([
            [value for _ in range(COLUMNS)]
            for row, value in zip(range(ROWS), VALUES)
        ])


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Simple benchmark for various Excel/XLSX python libraries'
    )
    parser.add_argument(
        '--filter', '-f',
        metavar='regex', type=re.compile,
        default=re.compile(''),
        help='benchmark filter',
    )

    parser.add_argument(
        '--tests', '-t',
        metavar='number', type=int,
        default=10,
        help='number of test runs',
    )

    parser.add_argument(
        '--columns', '-c',
        metavar='number', type=int,
        default=100,
        help='number of test spreadsheed columns',
    )

    parser.add_argument(
        '--rows', '-r',
        metavar='number', type=int,
        default=1000,
        help='number of test spreadsheet rows',
    )

    args = parser.parse_args()

    for stmt in sorted(filter(lambda fun: args.filter.search(fun.__name__),
                              get_benchmarks()),
                       key=lambda fun: fun.__name__):
        if hasattr(stmt, 'skip'):
            print("# SKIP {0} ({1})".format(stmt.__name__, stmt.skip))
            continue

        # do it globally to simplify test running and avoid
        # passing parameters to benchmark_function
        # (timeit - I hate you!)
        COLUMNS = args.columns
        ROWS = args.rows

        timer = timeit.Timer(stmt, 'gc.enable()')
        try:
            result = timer.timeit(number=args.tests)
        except ImportError as err:
            print("# SKIP {0} ({1})".format(stmt.__name__, err))
            continue

        print("{0:30} {1:5f}".format(stmt.__name__, result/args.tests,))
